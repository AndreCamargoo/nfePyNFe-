from django.conf import settings
from datetime import datetime
from django.http import HttpResponse

from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated, IsAdminUser

from rest_framework.response import Response

from .models import EventoCadastroEmpresa, EventoContato
from .serializer import (EventoCadastroEmpresaModelSerializer, EventoContatoModelSerializer)

from app.utils import utils

# Importações para PDF
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors


from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from rest_framework.parsers import MultiPartParser
from rest_framework.views import APIView
from openpyxl import load_workbook


class EventoCadastroListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = []
    serializer_class = EventoCadastroEmpresaModelSerializer
    queryset = EventoCadastroEmpresa.objects.all()
    pagination_class = utils.CustomPageSizePagination


class EventoCadastroRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = EventoCadastroEmpresaModelSerializer
    queryset = EventoCadastroEmpresa.objects.all()


class EventoContatoListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = []
    serializer_class = EventoContatoModelSerializer
    queryset = EventoContato.objects.all()
    pagination_class = utils.CustomPageSizePagination


class EventoContatoRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = EventoContatoModelSerializer
    queryset = EventoContato.objects.all()


class EventoImportXLSX(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        arquivo = request.FILES.get("arquivo")

        if not arquivo:
            return Response({"error": "Arquivo XLSX não enviado."}, status=400)

        try:
            wb = load_workbook(filename=arquivo)
            ws = wb.active
        except:
            return Response({"error": "Erro ao ler o arquivo XLSX."}, status=400)

        # Cabeçalhos esperados (exatamente igual o arquivo de exportação)
        headers_esperados = [
            "Nome Empresa", "Documento", "Endereço", "Cidade", "UF",
            "Email Empresa", "Telefone Empresa", "Status",
            "Nome Contato", "Cargo", "Email Contato", "Telefone Contato", "Origem Lead"
        ]

        headers_arquivo = [cell.value for cell in ws[1]]

        if headers_arquivo != headers_esperados:
            return Response({
                "error": "O arquivo XLSX não possui o layout esperado.",
                "esperado": headers_esperados,
                "recebido": headers_arquivo
            }, status=400)

        importados = []
        erros = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                (
                    nome_empresa, documento, endereco, cidade, uf,
                    email_empresa, telefone_empresa, status_empresa,
                    nome_contato, cargo_contato, email_contato,
                    telefone_contato, origem_lead
                ) = row

                # ==========================
                # IMPORTA CONTATO
                # ==========================
                contato, created_c = EventoContato.objects.get_or_create(
                    email=email_contato,
                    defaults={
                        "nome": nome_contato,
                        "cargo": cargo_contato,
                        "telefone": telefone_contato,
                        "origem_lead": origem_lead,
                        "status": "1"
                    }
                )

                # Atualiza caso exista
                if not created_c:
                    contato.nome = nome_contato
                    contato.cargo = cargo_contato
                    contato.telefone = telefone_contato
                    contato.origem_lead = origem_lead
                    contato.save()

                # ==========================
                # IMPORTA EMPRESA
                # ==========================
                empresa, created_e = EventoCadastroEmpresa.objects.update_or_create(
                    documento=documento,
                    defaults={
                        "nome_empresa": nome_empresa,
                        "endereco": endereco,
                        "cidade": cidade,
                        "uf": uf,
                        "email": email_empresa,
                        "telefone": telefone_empresa,
                        "status": "1" if status_empresa == "Ativo" else "2",
                        "contato": contato,
                    }
                )

                importados.append({
                    "linha": idx,
                    "empresa": empresa.nome_empresa,
                    "contato": contato.nome,
                    "created_empresa": created_e,
                    "created_contato": created_c
                })

            except Exception as e:
                erros.append({"linha": idx, "erro": str(e)})

        return Response({
            "message": "Importação concluída",
            "total_importados": len(importados),
            "total_erros": len(erros),
            "itens_importados": importados,
            "erros": erros
        })


class EventoDownload(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EventoCadastroEmpresaModelSerializer

    def post(self, request, *args, **kwargs):
        data_inicial = request.data.get('data_inicial')
        data_final = request.data.get('data_final')
        senha = request.data.get('senha')
        tipo = request.data.get('tipo', 'pdf').lower()

        # Verifica senha
        # if not senha or senha != settings.DOWNLOAD_AGENDA:
        #     return Response(
        #         {"error": "Senha inválida ou não informada."},
        #         status=status.HTTP_403_FORBIDDEN
        #     )

        # Valida datas
        if not data_inicial or not data_final:
            return Response(
                {"error": "Os campos 'data_inicial' e 'data_final' são obrigatórios."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            data_inicial = datetime.strptime(data_inicial, "%Y-%m-%d")
            data_final = datetime.strptime(data_final, "%Y-%m-%d")
        except ValueError:
            return Response(
                {"error": "As datas devem estar no formato YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Busca registros
        queryset = EventoCadastroEmpresa.objects.filter(
            created_at__date__gte=data_inicial,
            created_at__date__lte=data_final
        ).select_related('contato')

        if not queryset.exists():
            return Response(
                {"message": "Nenhum registro encontrado para o período informado."},
                status=status.HTTP_204_NO_CONTENT
            )

        serializer = self.get_serializer(queryset, many=True)

        # ===============================
        # GERAÇÃO DE PDF
        # ===============================
        if tipo == 'pdf':
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph("<b>Relatório de Cadastros de Empresas</b>", styles['Title']))
            elements.append(Paragraph(
                f"Período: {data_inicial.strftime('%d/%m/%Y')} a {data_final.strftime('%d/%m/%Y')}",
                styles['Normal']
            ))
            elements.append(Spacer(1, 20))

            for empresa in serializer.data:
                contato = empresa.get('contato')

                # Contato primeiro
                if contato:
                    elements.append(Paragraph("<b>Contato Vinculado</b>", styles['Heading3']))
                    table_data = [['Nome', 'Cargo', 'Email', 'Telefone', 'Origem']]
                    table_data.append([
                        contato['nome'],
                        contato['cargo'],
                        contato['email'],
                        contato['telefone'],
                        contato['origem_lead'],
                    ])
                    table = Table(table_data, colWidths=[100, 80, 130, 80, 70])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    elements.append(table)
                    elements.append(Spacer(1, 10))
                else:
                    elements.append(Paragraph("<i>Sem contato vinculado.</i>", styles['Italic']))
                    elements.append(Spacer(1, 10))

                # Empresa depois
                elements.append(Paragraph(
                    f"<b>{empresa['nome_empresa']}</b> ({empresa['documento']})",
                    styles['Heading3']
                ))
                elements.append(Paragraph(
                    f"Endereço: {empresa['endereco']}, {empresa['cidade']} - {empresa['uf']}",
                    styles['Normal']
                ))
                elements.append(Paragraph(
                    f"Email: {empresa['email']} | Telefone: {empresa['telefone']}",
                    styles['Normal']
                ))
                elements.append(Paragraph(
                    f"Status: {'Ativo' if empresa['status'] == '1' else 'Inativo'}",
                    styles['Normal']
                ))
                elements.append(Spacer(1, 20))

            doc.build(elements)
            pdf = buffer.getvalue()
            buffer.close()

            filename = f"relatorio_cadastros_{data_inicial.strftime('%Y%m%d')}_{data_final.strftime('%Y%m%d')}.pdf"
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # ===============================
        # GERAÇÃO DE XLSX
        # ===============================
        elif tipo == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = "Cadastros de Empresas"

            headers = [
                "Nome Empresa", "Documento", "Endereço", "Cidade", "UF",
                "Email Empresa", "Telefone Empresa", "Status",
                "Nome Contato", "Cargo", "Email Contato", "Telefone Contato", "Origem Lead"
            ]
            ws.append(headers)

            for empresa in serializer.data:
                contato = empresa.get('contato') or {}
                ws.append([
                    empresa['nome_empresa'],
                    empresa['documento'],
                    empresa['endereco'],
                    empresa['cidade'],
                    empresa['uf'],
                    empresa['email'],
                    empresa['telefone'],
                    'Ativo' if empresa['status'] == '1' else 'Inativo',
                    contato.get('nome', ''),
                    contato.get('cargo', ''),
                    contato.get('email', ''),
                    contato.get('telefone', ''),
                    contato.get('origem_lead', '')
                ])

            # Ajusta largura automática das colunas
            for col in ws.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f"relatorio_cadastros_{data_inicial.strftime('%Y%m%d')}_{data_final.strftime('%Y%m%d')}.xlsx"
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        # Caso formato não reconhecido
        else:
            return Response(
                {"error": "Tipo de arquivo inválido. Use 'pdf' ou 'xlsx'."},
                status=status.HTTP_400_BAD_REQUEST
            )
