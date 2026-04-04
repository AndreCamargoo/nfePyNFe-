# leads_api/services/import_service.py
import csv
import io
import re
import logging
from datetime import datetime
from django.db import transaction
from leads_api.models import Lead, Company, Product, Contact, Cnes
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ImportService:

    DEFAULT_COMPANIES = ["numb3rs"]
    DEFAULT_PRODUCTS = ["dbsaúde"]

    @staticmethod
    def get_default_companies():
        companies = []
        for nome in ImportService.DEFAULT_COMPANIES:
            company = Company.objects.filter(nome__iexact=nome).first()
            if company:
                companies.append(company)
            else:
                logger.warning(f"Empresa padrão '{nome}' não encontrada no banco.")
        return companies

    @staticmethod
    def get_default_products():
        products = []
        for nome in ImportService.DEFAULT_PRODUCTS:
            product = Product.objects.filter(nome__iexact=nome).first()
            if product:
                products.append(product)
            else:
                logger.warning(f"Produto padrão '{nome}' não encontrado no banco.")
        return products

    @staticmethod
    def normalize_string(value):
        if not value or not isinstance(value, str):
            return ''
        return re.sub(r'\s+', ' ', value.strip()).lower()

    @staticmethod
    def normalize_lower_string(value):
        return ImportService.normalize_string(value)

    @staticmethod
    def normalize_email(value):
        if not value or not isinstance(value, str):
            return ''
        return value.strip().lower()

    @staticmethod
    def clean_numeric(value):
        if not value:
            return ''
        cleaned = re.sub(r'[^0-9]', '', str(value))
        return cleaned

    @staticmethod
    def detect_file_type(file):
        filename = getattr(file, 'name', '')
        if filename.lower().endswith('.xlsx'):
            return 'xlsx'
        elif filename.lower().endswith('.xls'):
            return 'xls'
        elif filename.lower().endswith('.csv'):
            return 'csv'

        file.seek(0)
        header = file.read(4)
        file.seek(0)

        if header.startswith(b'PK'):
            return 'xlsx'
        elif header.startswith(b'\xD0\xCF\x11\xE0'):
            return 'xls'
        else:
            return 'csv'

    @staticmethod
    def read_xlsx_file(file):
        workbook = load_workbook(file, data_only=True)
        sheet = workbook.active

        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                headers.append('')

        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {}
            has_data = False
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    cell_value = str(value).strip() if value is not None else ''
                    if cell_value:
                        has_data = True
                    row_data[headers[idx]] = cell_value
            if has_data:
                rows.append((row_idx, row_data))

        return rows

    @staticmethod
    def read_csv_file(file):
        try:
            decoded_file = file.read().decode('utf-8')
        except UnicodeDecodeError:
            file.seek(0)
            decoded_file = file.read().decode('latin-1')

        if decoded_file.startswith('\ufeff'):
            decoded_file = decoded_file[1:]
        if decoded_file.startswith('ï»¿'):
            decoded_file = decoded_file[3:]

        io_string = io.StringIO(decoded_file)
        first_line = io_string.readline()
        io_string.seek(0)

        delimiter = ';' if ';' in first_line else ','
        reader = csv.DictReader(io_string, delimiter=delimiter)

        if not reader.fieldnames:
            raise Exception("Arquivo CSV vazio ou formato inválido")

        rows = []
        for idx, row in enumerate(reader, start=2):
            clean_row = {}
            has_data = False
            for key, value in row.items():
                if key is not None:
                    clean_key = str(key).strip()
                    clean_value = str(value).strip() if value else ''
                    clean_row[clean_key] = clean_value
                    if clean_value:
                        has_data = True
            if has_data:
                rows.append((idx, clean_row))

        return rows

    @staticmethod
    def process_csv(file, duplicate=False, celery=True):
        if celery:
            from leads_api.tasks import import_leads_csv_task

            file.seek(0)
            file_content = file.read()
            file_type = ImportService.detect_file_type(file)

            task = import_leads_csv_task.delay(file_content, file.name, duplicate, file_type)

            return {
                "task_id": task.id,
                "status": "processing",
                "message": "Importação iniciada em background",
                "file_type": file_type
            }

        return ImportService._process_file_sync(file, duplicate)

    @staticmethod
    def _process_file_sync(file, duplicate=False):
        file_type = ImportService.detect_file_type(file)

        if file_type in ['xlsx', 'xls']:
            rows = ImportService.read_xlsx_file(file)
        else:
            rows = ImportService.read_csv_file(file)

        results = {
            "created": 0,
            "updated": 0,
            "errors": [],
            "cnes_encontrados": 0,
            "cnes_nao_encontrados": 0,
            "invalid_rows": [],
            "success_rows": [],
            "file_type": file_type,
            "total_rows": len(rows)
        }

        cnes_cache = {}

        for row_num, row in rows:
            try:
                clean_row = {}
                for old_key, value in row.items():
                    if not old_key:
                        continue
                    clean_key = old_key.strip().replace('\ufeff', '').replace('ï»¿', '').strip()
                    clean_value = value.strip() if value else ""
                    clean_row[clean_key] = clean_value

                if clean_row.get('Razao Social', '').startswith('#'):
                    continue

                razao_social = clean_row.get('Razao Social', '')
                nome_conta = clean_row.get('Nome da conta', '')
                cnpj = clean_row.get('CNPJ', '')

                if not razao_social and not cnpj and not nome_conta:
                    results['invalid_rows'].append({
                        "linha": row_num,
                        "motivo": "Razão Social, Nome da conta e CNPJ não preenchidos",
                        "dados": clean_row
                    })
                    continue

                try:
                    with transaction.atomic():
                        created, updated = ImportService._process_row(clean_row, results, duplicate, cnes_cache)
                        if created:
                            results['created'] += 1
                        if updated:
                            results['updated'] += 1
                        results['success_rows'].append(row_num)
                except Exception as e:
                    error_msg = f"Linha {row_num}: {str(e)}"
                    results['errors'].append(error_msg)
                    logger.error(error_msg, exc_info=True)

            except Exception as e:
                error_msg = f"Linha {row_num}: Erro inesperado - {str(e)}"
                results['errors'].append(error_msg)
                logger.error(error_msg, exc_info=True)

        if results['invalid_rows']:
            report_path = ImportService._generate_error_report(results['invalid_rows'])
            results['error_report_path'] = report_path

        logger.info(f"Importação concluída: {results['created']} criados, {results['updated']} atualizados, {len(results['errors'])} erros")

        return results

    @staticmethod
    def _generate_error_report(invalid_rows):
        filename = f"importacao_erros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        filepath = f"reports/{filename}"

        import os
        os.makedirs('reports', exist_ok=True)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("RELATÓRIO DE LINHAS NÃO IMPORTADAS\n")
            f.write(f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
            f.write("=" * 80 + "\n\n")

            for row in invalid_rows:
                f.write(f"Linha {row['linha']}: {row['motivo']}\n")
                f.write(f"Dados: {row['dados']}\n")
                f.write("-" * 80 + "\n\n")

        return filepath

    @staticmethod
    def _enriquecer_com_cnes(row, cnes_cache):
        cnes_valor = row.get('CNES', '').strip()
        if not cnes_valor:
            return row, False

        if cnes_valor in cnes_cache:
            return cnes_cache[cnes_valor]['row'], cnes_cache[cnes_valor]['found']

        try:
            cnes_record = Cnes.objects.filter(cnes=cnes_valor).first()

            if not cnes_record:
                cnes_sem_zeros = cnes_valor.lstrip('0')
                if cnes_sem_zeros:
                    cnes_record = Cnes.objects.filter(cnes__icontains=cnes_sem_zeros).first()

            if cnes_record:
                logger.info(f"CNES {cnes_valor} encontrado! Sobrescrevendo dados...")

                if cnes_record.razao_social or cnes_record.fantasia:
                    row['Razao Social'] = ImportService.normalize_string(cnes_record.razao_social or cnes_record.fantasia)
                if cnes_record.cod_nat_jur:
                    row['Código Natureza Jurídica'] = cnes_record.cod_nat_jur
                if cnes_record.natureza_juridica:
                    row['Natureza Jurídica'] = ImportService.normalize_string(cnes_record.natureza_juridica)
                if cnes_record.cpf_cnpj:
                    row['CNPJ'] = ImportService.clean_numeric(cnes_record.cpf_cnpj)
                if cnes_record.telefone:
                    row['Conta: Telefone'] = ImportService.clean_numeric(cnes_record.telefone)
                if cnes_record.cidade:
                    row['Cidade de correspondência'] = ImportService.normalize_string(cnes_record.cidade)
                if cnes_record.uf:
                    row['Estado/Província de correspondência'] = cnes_record.uf.lower()

                cnes_cache[cnes_valor] = {'row': row, 'found': True}
                return row, True
            else:
                cnes_cache[cnes_valor] = {'row': row, 'found': False}
                return row, False
        except Exception as e:
            logger.error(f"Erro ao buscar CNES {cnes_valor}: {str(e)}")
            return row, False

    @staticmethod
    def _parse_contacts_from_row(row):
        contacts = []

        primeiro_nome = row.get('Primeiro Nome', '')
        sobrenome = row.get('Sobrenome', '')
        cargo = row.get('Cargo', '')
        email = row.get('Email', '')
        celular = row.get('Celular', '')
        telefone_contato = row.get('Telefone', '')
        email_extra = row.get('Email Secundário', '')

        primeiro_nome = ImportService.normalize_string(primeiro_nome) if primeiro_nome else ''
        sobrenome = ImportService.normalize_string(sobrenome) if sobrenome else ''
        cargo = ImportService.normalize_string(cargo) if cargo else ''
        email = ImportService.normalize_email(email) if email else ''
        celular = ImportService.clean_numeric(celular) if celular else ''
        telefone_contato = ImportService.clean_numeric(telefone_contato) if telefone_contato else ''
        email_extra = ImportService.normalize_email(email_extra) if email_extra else ''

        if primeiro_nome or sobrenome or email or celular or telefone_contato:
            nome_completo = ""
            if primeiro_nome and sobrenome:
                nome_completo = f"{primeiro_nome} {sobrenome}".strip()
            elif primeiro_nome:
                nome_completo = primeiro_nome
            elif sobrenome:
                nome_completo = sobrenome

            if not nome_completo and email:
                nome_completo = email.split('@')[0].lower()

            if nome_completo:
                contact = {
                    'nome': nome_completo[:500],
                    'setor': cargo[:300] if cargo else '',
                    'email': email[:500] if email else '',
                    'celular': celular[:100] if celular else '',
                    'telefone_contato': telefone_contato[:100] if telefone_contato else '',
                    'email_extra': email_extra[:500] if email_extra else ''
                }
                contacts.append(contact)
        return contacts

    @staticmethod
    def _find_existing_lead(row):
        """Busca lead existente por CNPJ, CNES, Razao Social ou Nome da conta (Busca Cruzada)"""
        cnpj_raw = row.get('CNPJ', '')
        cnes_raw = row.get('CNES', '')
        razao_social = ImportService.normalize_string(row.get('Razao Social', ''))
        nome_conta = ImportService.normalize_string(row.get('Nome da conta', ''))

        # 1. Busca Exata por CNPJ
        if cnpj_raw:
            cnpj_limpo = ImportService.clean_numeric(cnpj_raw)
            if cnpj_limpo:
                lead = Lead.objects.filter(cnpj=cnpj_limpo).first()
                if lead: 
                    return lead

        # 2. Busca Exata por CNES (Identificador forte da área de saúde)
        if cnes_raw:
            cnes_limpo = str(cnes_raw).strip()
            if cnes_limpo:
                lead = Lead.objects.filter(cnes=cnes_limpo).first()
                if lead:
                    return lead

        # 3. Busca Cruzada de Nomes
        # Agrupamos todos os nomes válidos fornecidos na planilha para este registro
        nomes_para_testar = []
        if razao_social:
            nomes_para_testar.append(razao_social)
        if nome_conta:
            nomes_para_testar.append(nome_conta)

        for nome in nomes_para_testar:
            # Procura se esse nome bate com a 'empresa' de algum registro
            lead = Lead.objects.filter(empresa__iexact=nome).first()
            if lead:
                return lead

            # Procura se esse nome bate com o 'apelido' de algum registro
            lead = Lead.objects.filter(apelido__iexact=nome).first()
            if lead:
                return lead

        return None

    @staticmethod
    def _update_lead_fields(lead, row, cnes_encontrado):
        updated = False

        razao_social = ImportService.normalize_string(row.get('Razao Social', ''))
        if razao_social and lead.empresa != razao_social:
            lead.empresa = razao_social[:500]
            updated = True

        nome_conta = ImportService.normalize_string(row.get('Nome da conta', ''))
        if nome_conta and lead.apelido != nome_conta:
            lead.apelido = nome_conta[:500]
            updated = True

        cnpj_raw = row.get('CNPJ', '')
        if cnpj_raw:
            cnpj_limpo = ImportService.clean_numeric(cnpj_raw)
            if cnpj_limpo and lead.cnpj != cnpj_limpo:
                lead.cnpj = cnpj_limpo[:50]
                updated = True

        if cnes_encontrado:
            cnes_valor = row.get('CNES', '').strip()
            if cnes_valor and lead.cnes != cnes_valor:
                lead.cnes = cnes_valor[:50]
                updated = True
        elif row.get('CNES'):
            cnes_valor = row.get('CNES', '').strip()
            if cnes_valor and lead.cnes != cnes_valor:
                lead.cnes = cnes_valor[:50]
                updated = True

        telefone = row.get('Conta: Telefone', '')
        if telefone:
            telefone_limpo = ImportService.clean_numeric(telefone)
            if telefone_limpo and lead.telefone != telefone_limpo:
                lead.telefone = telefone_limpo[:100]
                updated = True

        cidade = row.get('Cidade de correspondência', '')
        if cidade:
            cidade_norm = ImportService.normalize_string(cidade)
            if cidade_norm and lead.cidade != cidade_norm:
                lead.cidade = cidade_norm[:200]
                updated = True

        estado = row.get('Estado/Província de correspondência', '')
        if estado:
            estado_norm = estado.lower().strip()[:2]
            if estado_norm and lead.estado != estado_norm:
                lead.estado = estado_norm
                updated = True

        # --- BLOCO: Classificação (É Cliente?) GARANTIDO ---
        chave_cliente = None
        for k in row.keys():
            if k and 'cliente' in str(k).lower():
                chave_cliente = k
                break

        e_cliente_raw = str(row.get(chave_cliente, '')).strip().lower() if chave_cliente else ''

        if e_cliente_raw in ['verdadeiro', 'true', 'v', '1', 't', 'sim', 's', 'y', 'yes']:
            classificacao = 'Cliente'
        else:
            classificacao = 'Não Cliente'

        classificacao_banco = (lead.classificacao or '').lower()
        if classificacao_banco != classificacao.lower():
            lead.classificacao = classificacao
            updated = True
        # ----------------------------------------------------

        origem = row.get('Origem', '')
        if origem:
            origem_norm = ImportService.normalize_string(origem)
            if origem_norm and lead.origem != origem_norm:
                lead.origem = origem_norm[:200]
                updated = True

        cod_nat_jur = row.get('Código Natureza Jurídica', '')
        if cod_nat_jur and lead.cod_nat_jur != cod_nat_jur:
            lead.cod_nat_jur = cod_nat_jur[:50]
            updated = True

        natureza_jur = row.get('Natureza Jurídica', '')
        if natureza_jur:
            natureza_norm = ImportService.normalize_string(natureza_jur)
            if natureza_norm and lead.natureza_juridica != natureza_norm:
                lead.natureza_juridica = natureza_norm[:500]
                updated = True

        segmento = row.get('Segmento', '')
        if segmento:
            segmento_norm = ImportService.normalize_string(segmento)
            if segmento_norm and lead.segmento != segmento_norm:
                lead.segmento = segmento_norm[:200]
                updated = True

        if updated:
            lead.save()
            logger.info(f"Lead salvo/atualizado: {lead.empresa}")

        return updated

    @staticmethod
    def _process_empresas_grupo(lead, row):
        grupos_str = row.get('Grupo Empresa', '')
        lead.empresas_grupo.clear()

        if not grupos_str or grupos_str.strip() == '':
            default_companies = ImportService.get_default_companies()
            for company in default_companies:
                lead.empresas_grupo.add(company)
            return

        empresas_adicionadas = 0
        for nome in grupos_str.split(','):
            nome_limpo = ImportService.normalize_string(nome)
            if nome_limpo:
                company = Company.objects.filter(nome__iexact=nome_limpo[:255]).first()
                if company:
                    lead.empresas_grupo.add(company)
                    empresas_adicionadas += 1

        if empresas_adicionadas == 0:
            default_companies = ImportService.get_default_companies()
            for company in default_companies:
                lead.empresas_grupo.add(company)

    @staticmethod
    def _process_produtos_interesse(lead, row):
        produtos_str = row.get('Produtos Interesse', '')
        lead.produtos_interesse.clear()

        if not produtos_str or produtos_str.strip() == '':
            default_products = ImportService.get_default_products()
            for product in default_products:
                lead.produtos_interesse.add(product)
            return

        produtos_adicionados = 0
        for nome in produtos_str.split(','):
            nome_limpo = ImportService.normalize_string(nome)
            if nome_limpo:
                product = Product.objects.filter(nome__iexact=nome_limpo[:255]).first()
                if product:
                    lead.produtos_interesse.add(product)
                    produtos_adicionados += 1

        if produtos_adicionados == 0:
            default_products = ImportService.get_default_products()
            for product in default_products:
                lead.produtos_interesse.add(product)

    @staticmethod
    def _process_contact(lead, contact_data):
        email = contact_data.get('email', '')

        if not email:
            Contact.objects.create(
                lead=lead,
                nome=contact_data.get('nome', '')[:500],
                setor=contact_data.get('setor', '')[:300],
                email='',
                celular=contact_data.get('celular', '')[:100],
                telefone_contato=contact_data.get('telefone_contato', '')[:100],
                email_extra=contact_data.get('email_extra', '')[:500]
            )
            return

        existing_contact = Contact.objects.filter(lead=lead, email=email).first()

        if existing_contact:
            updated = False
            if contact_data.get('nome') and existing_contact.nome != contact_data['nome']:
                existing_contact.nome = contact_data['nome']
                updated = True
            if contact_data.get('setor') and existing_contact.setor != contact_data['setor']:
                existing_contact.setor = contact_data['setor']
                updated = True
            if contact_data.get('celular') and existing_contact.celular != contact_data['celular']:
                existing_contact.celular = contact_data['celular']
                updated = True
            if contact_data.get('telefone_contato') and existing_contact.telefone_contato != contact_data['telefone_contato']:
                existing_contact.telefone_contato = contact_data['telefone_contato']
                updated = True
            if contact_data.get('email_extra') and existing_contact.email_extra != contact_data['email_extra']:
                existing_contact.email_extra = contact_data['email_extra']
                updated = True

            if updated:
                existing_contact.save()
        else:
            Contact.objects.create(
                lead=lead,
                nome=contact_data.get('nome', '')[:500],
                setor=contact_data.get('setor', '')[:300],
                email=email[:500],
                celular=contact_data.get('celular', '')[:100],
                telefone_contato=contact_data.get('telefone_contato', '')[:100],
                email_extra=contact_data.get('email_extra', '')[:500]
            )

    @staticmethod
    def _process_row(row, results, duplicate, cnes_cache):
        row_enriquecida, cnes_encontrado = ImportService._enriquecer_com_cnes(row, cnes_cache)
        if cnes_encontrado:
            results['cnes_encontrados'] += 1
        elif row.get('CNES'):
            results['cnes_nao_encontrados'] += 1

        row = row_enriquecida

        lead = ImportService._find_existing_lead(row)
        is_new = False
        is_updated = False

        if not lead:
            razao_social = ImportService.normalize_string(row.get('Razao Social', ''))
            nome_conta = ImportService.normalize_string(row.get('Nome da conta', ''))

            lead = Lead(empresa=razao_social[:500], apelido=nome_conta[:500])
            is_new = True

            ImportService._update_lead_fields(lead, row, cnes_encontrado)
            lead.save()
            logger.info(f"Criando novo lead com ID: {lead.id}")

            ImportService._process_empresas_grupo(lead, row)
            ImportService._process_produtos_interesse(lead, row)
        else:
            logger.info(f"Lead existente encontrado. Atualizando dados: {lead.empresa or lead.apelido}")
            if hasattr(lead, 'deleted_at') and lead.deleted_at:
                lead.deleted_at = None
                lead.save()

            is_updated = ImportService._update_lead_fields(lead, row, cnes_encontrado)

            ImportService._process_empresas_grupo(lead, row)
            ImportService._process_produtos_interesse(lead, row)

        contacts = ImportService._parse_contacts_from_row(row)
        for contact_data in contacts:
            ImportService._process_contact(lead, contact_data)

        return is_new, is_updated
